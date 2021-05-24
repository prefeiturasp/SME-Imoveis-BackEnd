from django.core.management.base import BaseCommand
from sme_ofertaimoveis.imovel.models import *
from sme_ofertaimoveis.dados_comuns.models import *

from openpyxl import Workbook, load_workbook
from openpyxl.writer.excel import save_virtual_workbook
from django.utils import timezone
import pytz
import datetime
import pdb
class Command(BaseCommand):
    help = "Corrigir logs de envio de email."

    def handle(self, *args, **kwrgs):
        self.stdout.write("Lendo dados da planilha.")
        # Trocar para nome da planilha correto
        self.fix_log_envio_email('Protocolos_Sistema Cadastro de Imóveis 12-05-2021.xlsx')
        self.stdout.write("Processo de atualização finalizado.")

    def fix_log_envio_email(self, nome_arquivo):
        wb = load_workbook(filename = nome_arquivo)
        # Trocar para sheet correto
        sheet_ranges = wb['Relatório por Status']
        linhas = sheet_ranges.max_row
        colunas = sheet_ranges.max_column
        for i in range(2, linhas + 1):
            protocolo = sheet_ranges.cell(row=i, column=3).value
            data_vistoria = sheet_ranges.cell(row=i, column=16).value
            data_retorno_email = sheet_ranges.cell(row=i, column=18).value
            tipo_retorno = sheet_ranges.cell(row=i, column=19).value
            tipo_status = sheet_ranges.cell(row=i, column=23).value
            imovel = Imovel.objects.get(id=protocolo)
            for idx, log in enumerate(imovel.logs.all()):
                print(f"====================PROTOCOLO: {protocolo}, LOG_INDEX: {idx}========================\n\n")
                if tipo_status in ["Demanda Insuficiente", "Cancelado", "Área insuficiente"]:
                    if idx != 1:
                        log.criado_em = imovel.criado_em
                        log.email_enviado = False
                    else:
                        if data_retorno_email != None:
                            log.criado_em = data_retorno_email
                            log.email_enviado = True
                        else:
                            log.criado_em = imovel.criado_em
                            log.email_enviado = False
                elif tipo_status in ["Reprovado", "Aprovado"]:
                    if idx < 7:
                        log.criado_em = imovel.criado_em
                        log.email_enviado = False
                    else:
                        if data_retorno_email != None:
                            log.criado_em = data_retorno_email
                            log.email_enviado = True if (idx == 7) else False
                        else:
                            log.criado_em = data_vistoria if (data_vistoria != None) else imovel.criado_em
                            log.email_enviado = False
                else:
                    pass
                log.save()
