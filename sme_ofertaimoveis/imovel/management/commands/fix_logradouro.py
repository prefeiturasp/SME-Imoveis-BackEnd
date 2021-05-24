from django.core.management.base import BaseCommand
from sme_ofertaimoveis.imovel.models import Imovel
from openpyxl import Workbook, load_workbook
from openpyxl.writer.excel import save_virtual_workbook

class Command(BaseCommand):
    help = "Atualizar campo endereco dos imóveis cadastrados com base na planilha."

    def handle(self, *args, **kwrgs):
        self.stdout.write("Lendo dados da planilha.")
        # Trocar para nome da planilha correto
        self.fix_logradouro('planilha_com_correcao_enderecos.xlsx')
        self.stdout.write("Processo de atualização finalizado.")

    def fix_logradouro(self, nome_arquivo):
        wb = load_workbook(filename = nome_arquivo)
        # Trocar para sheet correto
        sheet_ranges = wb['Sheet']
        linhas = sheet_ranges.max_row
        colunas = sheet_ranges.max_column
        for i in range(2, linhas + 1):
            protocolo = sheet_ranges.cell(row=i, column=1).value
            imovel = Imovel.objects.get(id=protocolo)
            imovel.endereco =  sheet_ranges.cell(row=i, column=3).value
            imovel.save()
