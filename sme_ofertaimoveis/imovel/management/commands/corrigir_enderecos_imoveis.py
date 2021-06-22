from django.core.management.base import BaseCommand
from openpyxl import load_workbook
import requests
import time
from sme_ofertaimoveis.imovel.models import Imovel
from sme_ofertaimoveis.dados_comuns.models import Setor

class Command(BaseCommand):
    help = "Exportar planilha com informações do endereço com base no CEP."

    def handle(self, *args, **kwrgs):
        self.stdout.write("Iniciando processo de correcao dos endereços.")
        self.stdout.write("Lendo dados da planilha.")
        self.atualiza_imovel('planilha_enderecos_atualizados.xlsx')

    def atualiza_imovel(self, nome_arquivo):
        wb = load_workbook(filename = nome_arquivo)
        # Trocar para sheet correto
        sheet_ranges = wb['Relatório endereço']
        linhas = sheet_ranges.max_row
        colunas = sheet_ranges.max_column
        for i in range(2, linhas + 1):
            imovel = Imovel.objects.get(id=sheet_ranges.cell(row=i, column=1).value)
            print('\n\n++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++\n\n')
            print(f"Atualizando informações com base na planilha do protocolo {imovel.id}.")
            imovel.cep = sheet_ranges.cell(row=i, column=2).value
            imovel.endereco = sheet_ranges.cell(row=i, column=3).value
            imovel.numero = sheet_ranges.cell(row=i, column=4).value
            imovel.bairro = sheet_ranges.cell(row=i, column=5).value
            endereco = imovel.endereco + " " + imovel.numero
            self.atualiza_latitude_longitude(imovel, endereco)

    def atualiza_latitude_longitude(self, imovel, endereco):
        try:
            print(f"Atualizando latitude e longitude do protocolo {imovel.id}.")
            response = requests.get(f"https://georef.sme.prefeitura.sp.gov.br/v1/search?text={endereco}&layers=address&boundary.gid=whosonfirst:locality:101965533")
            result = response.json()
            imovel.latitude = result['features'][0]['geometry']['coordinates'][1]
            imovel.longitude = result['features'][0]['geometry']['coordinates'][0]
            self.atualiza_setor(imovel, imovel.latitude, imovel.longitude)
            imovel.save()
            time.sleep(5)
        except Exception as e:
            print(f"Erro ao atualizar cadastro: {imovel.protocolo}")
            print(f"Erro: {e}")

    def atualiza_setor(self, imovel, latitude, longitude):
        print(f"Atualizando setor do protocolo {imovel.id}.")
        response = requests.get(f"https://escolaaberta.sme.prefeitura.sp.gov.br/api/localizador?lat={latitude}&lon={longitude}&radius=100")
        result = response.json()
        setor_api = str(result['results'][0]['setor'])
        if len(setor_api) < 4:
            setor = setor_api
            for x in range(4 - len(setor_api)):
                setor = '0' + setor
            imovel.setor = Setor.objects.filter(codigo=setor).first()
        else:
            imovel.setor = Setor.objects.filter(codigo=setor_api).first()
