import re

from openpyxl import Workbook
from openpyxl.writer.excel import save_virtual_workbook
from django.core.management.base import BaseCommand
from sme_ofertaimoveis.imovel.models import Imovel

class Command(BaseCommand):
    help = "Atualizar campo logradouro com base na api de consulta de CEP."

    def handle(self, *args, **kwrgs):
        self.stdout.write("Exportando planilha.")

        cabecalho = ['Protocolo', 'Endereço', 'Correção' ]
        wb = Workbook()
        ws = wb.active
        for ind, nome_coluna in enumerate(cabecalho, 1):
            celula = ws.cell(row=1, column=ind)
            celula.value = nome_coluna
            ws.column_dimensions[celula.column_letter].width = 10 if ind == 1 else 50

        for indx, imovel in enumerate(Imovel.objects.all(), 2):
            if imovel.id != 2544:
                endereco = re.split('\d+', imovel.endereco)[0].split(',')[0]
                ws.cell(row=indx, column=1, value=imovel.protocolo)
                ws.cell(row=indx, column=2, value=imovel.endereco)
                ws.cell(row=indx, column=3, value=endereco)

        wb.save("planilha_com_correcao_enderecos.xlsx")
        self.stdout.write("Processo de exportação finalizado.")
