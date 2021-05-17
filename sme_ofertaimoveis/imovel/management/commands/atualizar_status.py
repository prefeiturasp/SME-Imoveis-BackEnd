from django.core.management.base import BaseCommand
from sme_ofertaimoveis.imovel.models import Imovel, Proponente
from sme_ofertaimoveis.users.models import User
from openpyxl import load_workbook
import datetime

class Command(BaseCommand):
    help = "Atualizar cadastros da base de dados de produção com base na Planilha atualizada."

    def handle(self, *args, **kwrgs):
        self.stdout.write("Atualizando base de dados.")
        # Trocar para nome da planilha correto
        dados_planilha = self.get_dados_planilha('Protocolos_Sistema Cadastro de Imóveis 12-05-2021.xlsx')
        self.atualizar_status(dados_planilha)

        self.stdout.write("Processo de atualização finalizado.")

    def get_dados_planilha(self, nome_arquivo):
        wb = load_workbook(filename = nome_arquivo)
        # Trocar para sheet correto
        sheet_ranges = wb['Relatório por Status']
        linhas = sheet_ranges.max_row
        colunas = sheet_ranges.max_column
        dados_planilha = []

        for i in range(2, linhas + 1):
            imovel = {}
            for j in range(1, colunas + 1):
                imovel[sheet_ranges.cell(row=1, column=j).value] = sheet_ranges.cell(row=i, column=j).value
            dados_planilha.append(imovel)

        return dados_planilha

    def atualizar_status(self, dados_planilha):
        # Trocar para o Usuário correto.
        admin = User.objects.get(username="XXXXXXX")
        data_cancelamento = datetime.datetime.now()

        for dado in dados_planilha:
            print(f"======================================================== Atualizando Cadastro {dado['Protocolo']} ========================================================\n\n")
            imovel = Imovel.objects.get(id=dado['Protocolo'])
            self.atualizar_dados(dado, imovel)
            if dado['Tipo Status'] == "Aprovado":
                self.criar_logs_anteriores(dado, imovel, admin)
                imovel.aprova_vistoria(user=admin, enviar_email=False)
                imovel.envia_a_dre(user=admin, enviar_email=False,
                                   processo_sei=dado['Processo SEI'], nome_da_unidade=dado['Nome da Unidade'])
                imovel.finaliza_aprovado(user=admin, enviar_email=False)
            elif dado['Tipo Status'] == "Demanda Insuficiente":
                imovel.sme_analisa_previamente(user=admin, enviar_email=False)
                imovel.finaliza_demanda_insuficiente(user=admin, enviar_email=False)
            elif dado['Tipo Status'] == "Área insuficiente":
                imovel.sme_analisa_previamente(user=admin, enviar_email=False)
                imovel.finaliza_area_insuficiente(user=admin, enviar_email=False)
            elif dado['Tipo Status'] == "Reprovado":
                self.criar_logs_anteriores(dado, imovel, admin)
                imovel.reprova_vistoria(user=admin, enviar_email=False)
                imovel.finaliza_reprovado(user=admin, enviar_email=False)
            else:
                imovel.cancela(user=admin, data_agendada=data_cancelamento)
            imovel.save()
            print(f"======================================================== Cadastro {dado['Protocolo']} Atualizado ========================================================\n\n")

    def atualizar_dados(self, dado, imovel):
        imovel.status = Imovel.workflow_class.SOLICITACAO_REALIZADA
        imovel.endereco = dado['Endereço']
        imovel.numero = dado['Número']
        imovel.complemento = dado['Complemento']
        imovel.bairro = dado['Bairro']
        imovel.cidade = dado['Cidade']
        imovel.uf = dado['UF']
        imovel.cep = dado['Cep']
        imovel.latitude = dado['Latitude']
        imovel.longitude = dado['Longitude']
        if imovel.proponente:
            imovel.proponente.nome = dado['Nome do Proprietário']
        else:
            imovel.proponente = Proponente.objects.create(nome=dado['Nome do Proprietário'])
        if dado['Número do IPTU']:
            imovel.nao_possui_iptu = False
            imovel.numero_iptu = dado['Número do IPTU']
        else:
            imovel.nao_possui_iptu = True
            imovel.observacoes = "Migrado do imóveis 1.0"
        if dado['Área Construída (m²)']:
            area = dado['Área Construída (m²)']
            if type(area) != int:
                area = [int(s) for s in area if s.isdigit() and s != '²']
                area_formatada = ''.join([str(elem) for elem in area])
                imovel.area_construida = int(area_formatada)
            else:
                imovel.area_construida = int(area)

    def criar_logs_anteriores(self, dado, imovel, admin):
        imovel.sme_analisa_previamente(user=admin, enviar_email=False)
        imovel.envia_a_comapre(user=admin, enviar_email=False)
        imovel.agenda_vistoria(user=admin, data_agendada=dado['Data da vistoria'], enviar_email=False)
        imovel.aguarda_relatorio_vistoria(user=admin, enviar_email=False)
        imovel.relatorio_vistoria(user=admin, enviar_email=False)
        imovel.aguarda_laudo_valor_locaticio(user=admin, enviar_email=False)
        imovel.laudo_valor_locaticio(user=admin, enviar_email=False)
