from django.core.management.base import BaseCommand
from sme_ofertaimoveis.imovel.models import Imovel
from openpyxl import Workbook, load_workbook
from openpyxl.writer.excel import save_virtual_workbook

class Command(BaseCommand):
    help = "Atualizar campo numero_iptu dos imóveis cadastrados com base na planilha."

    def handle(self, *args, **kwrgs):
        self.stdout.write("Lendo dados da planilha.")
        # Trocar para nome da planilha correto
        self.fix_iptu_com_erro('Planilha IPTU divergentes 18-05-21.xlsx')
        self.add_mascara_para_iptu()
        self.texto_correto_observacoes()
        self.stdout.write("Processo de atualização finalizado.")

    def fix_iptu_com_erro(self, nome_arquivo):
        wb = load_workbook(filename = nome_arquivo)
        # Trocar para sheet correto
        sheet_ranges = wb['Planilha1']
        linhas = sheet_ranges.max_row
        colunas = sheet_ranges.max_column

        for i in range(2, linhas + 1):
            protocolo = sheet_ranges.cell(row=i, column=3).value
            imovel = Imovel.objects.get(id=protocolo)
            imovel.numero_iptu = ''
            imovel.nao_possui_iptu = True
            imovel.observacoes = "Cadastro oriundo da base Imóveis 1.0"
            imovel.save()

    def add_mascara_para_iptu(self):
        for imovel in Imovel.objects.all():
            if len(imovel.numero_iptu) == 11:
                imovel.numero_iptu = '{}{}{}.{}{}{}.{}{}{}{}.{}'.format(*imovel.numero_iptu)
            if len(imovel.numero_iptu) == "":
                imovel.nao_possui_iptu = True
                imovel.observacoes = "Cadastro oriundo da base Imóveis 1.0"
            imovel.save()

    def texto_correto_observacoes(self):
        for imovel in Imovel.objects.filter(observacoes="Migrado do imóveis 1.0"):
            imovel.observacoes = "Cadastro oriundo da base Imóveis 1.0"
            imovel.save()
