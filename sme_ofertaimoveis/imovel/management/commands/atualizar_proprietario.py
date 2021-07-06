from django.core.management.base import BaseCommand
from sme_ofertaimoveis.imovel.models import Imovel, Proponente, ContatoImovel
from openpyxl import Workbook, load_workbook
from sme_ofertaimoveis.users.models import User
import time

class Command(BaseCommand):
    help = "Atualizar informações do proprietário dos imóveis cadastrados com base na planilha."

    def handle(self, *args, **kwrgs):
        self.stdout.write("Atualizando planilha.")
        # Trocar para nome da planilha correto
        self.atualizar_informacoes('Planilha de Imóveis - Dados do Proprietário.xlsx')
        self.stdout.write("Processo de atualização finalizado.")

    def atualizar_informacoes(self, nome_arquivo):
        wb = load_workbook(filename = nome_arquivo)
        # Trocar para sheet correto
        sheet_ranges = wb['Query result']
        linhas = sheet_ranges.max_row
        colunas = sheet_ranges.max_column

        for i in range(2, linhas + 1):
            protocolo = sheet_ranges.cell(row=i, column=1).value
            self.stdout.write(f"+++++++++++++++ Atualizando cadastro {protocolo}. +++++++++++++++")

            try:
                imovel = Imovel.objects.get(id=protocolo)
                self.atualiza_proponente(imovel, sheet_ranges, i)
                self.atualiza_proprietario(imovel, sheet_ranges, i)
                imovel.save()
            except Exception as e:
                print(f"Erro ao atualizar cadastro: {protocolo}")
                print(f"Erro: {e}")
                print('cadastro não encontrado na base de produção')
                time.sleep(5)
                pass

    def atualiza_proponente(self, imovel, sheet_ranges, i):
        # Proponente
        try:
            imovel.proponente.cpf_cnpj = sheet_ranges.cell(row=i, column=2).value
            imovel.proponente.email = sheet_ranges.cell(row=i, column=3).value
            imovel.proponente.nome = sheet_ranges.cell(row=i, column=5).value
            imovel.proponente.telefone = sheet_ranges.cell(row=i, column=6).value
            imovel.proponente.tipo_proponente = 1

        except AttributeError:
            imovel.proponente = Proponente.objects.create(
                cpf_cnpj=sheet_ranges.cell(row=i, column=2).value,
                email=sheet_ranges.cell(row=i, column=3).value,
                nome=sheet_ranges.cell(row=i, column=5).value,
                telefone=sheet_ranges.cell(row=i, column=6).value,
                tipo_proponente = 1
            )

    def atualiza_proprietario(self, imovel, sheet_ranges, i):
        # Contato
        try:
            imovel.contato.cpf_cnpj = sheet_ranges.cell(row=i, column=2).value
            imovel.contato.email = sheet_ranges.cell(row=i, column=3).value
            imovel.contato.nome = sheet_ranges.cell(row=i, column=5).value
            imovel.contato.telefone = sheet_ranges.cell(row=i, column=6).value

        except AttributeError:
            imovel.contato = ContatoImovel.objects.create(
                cpf_cnpj=sheet_ranges.cell(row=i, column=2).value,
                email=sheet_ranges.cell(row=i, column=3).value,
                nome=sheet_ranges.cell(row=i, column=5).value,
                telefone=sheet_ranges.cell(row=i, column=6).value,
            )
