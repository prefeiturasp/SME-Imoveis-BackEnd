from django.core.management.base import BaseCommand
from openpyxl import Workbook
from openpyxl.writer.excel import save_virtual_workbook
from openpyxl.styles import PatternFill, Border, Side, Alignment, Protection, Font
from django.db import connection

from sme_ofertaimoveis.imovel.models import Imovel


class Command(BaseCommand):
    help = "Exportar planilha com informações dos imóveis duplicados."

    def handle(self, *args, **kwrgs):
        self.stdout.write("Exportando planilha.")

        cabecalho = ['Situação Duplicação', 'Excluído', 'Protocolo', 'Número do IPTU',
                     'Bairro', 'Número', 'Endereço', 'Complemento',
                     'Latitude', 'Longitude', 'Nome Proprietário',
                     'Cep', 'Cidade', 'UF' ]

        wb = Workbook()
        ws = wb.active
        ws.title = "Relatório por Status"
        cursor = connection.cursor()

        for ind, nome_coluna in enumerate(cabecalho, 1):
            celula = ws.cell(row=1, column=ind)
            celula.value = nome_coluna
            ws.column_dimensions[celula.column_letter].width = 24

        imoveis = Imovel.objects.all()
        for indx, imovel in enumerate(imoveis, 2):
            # Pegar nome contato diretamente do Banco de Dados
            raw_query = f"SELECT nome FROM imovel_contatoimovel WHERE id={imovel.id}"
            cursor.execute(raw_query)
            nome = cursor.fetchall()[0][0]

            ws.cell(row=indx, column=1, value=imovel.situacao_duplicidade)
            ws.cell(row=indx, column=2, value='Excluir' if imovel.excluido else 'Manter')
            ws.cell(row=indx, column=3, value=imovel.protocolo)
            ws.cell(row=indx, column=4, value=imovel.numero_iptu)
            ws.cell(row=indx, column=5, value=imovel.bairro)
            ws.cell(row=indx, column=6, value=imovel.numero)
            ws.cell(row=indx, column=7, value=imovel.endereco)
            ws.cell(row=indx, column=8, value=imovel.complemento)
            ws.cell(row=indx, column=9, value=imovel.latitude)
            ws.cell(row=indx, column=10, value=imovel.longitude)
            ws.cell(row=indx, column=11, value=nome)
            ws.cell(row=indx, column=12, value=imovel.cep)
            ws.cell(row=indx, column=13, value=imovel.cidade)
            ws.cell(row=indx, column=14, value=imovel.uf)

        wb.save("planilha_duplicadas.xlsx")

        self.stdout.write("Processo de exportação finalizado.")
