import datetime
import requests
import environ
from io import BytesIO

from openpyxl import Workbook
from openpyxl.writer.excel import save_virtual_workbook
from openpyxl.styles import PatternFill, Border, Side, Alignment, Protection, Font

from django.http import HttpResponse
from django.conf import settings

from django.template.loader import render_to_string
from django.core.files.storage import FileSystemStorage
from weasyprint import HTML

from rest_framework import status, mixins, viewsets
from rest_framework.decorators import action
from rest_framework.exceptions import ValidationError
from rest_framework.pagination import LimitOffsetPagination
from rest_framework.permissions import AllowAny, IsAuthenticated
from rest_framework.response import Response
from rest_framework.viewsets import ViewSet
from rest_framework.views import APIView
from ..models import Imovel, PlantaFoto, Proponente
from sme_ofertaimoveis.dados_comuns.models import DiretoriaRegional, Distrito, Setor
from .serializers import CadastroImovelSerializer, UpdateImovelSerializer, ListaImoveisSeriliazer, AnexoCreateSerializer, AnexoSerializer
from ..tasks import task_send_email_to_usuario, task_send_email_to_sme
from ..utils import checa_digito_verificador_iptu
from django.db.models import Q
from django.db.models import Sum
from itertools import chain

env = environ.Env()

class CadastroImoveisViewSet(viewsets.ModelViewSet,
                             mixins.CreateModelMixin,
                             mixins.UpdateModelMixin,
                             mixins.ListModelMixin):
    permission_classes = (AllowAny,)
    queryset = Imovel.objects.filter(excluido=False).all()
    serializer_class = ListaImoveisSeriliazer

    def get_serializer_class(self):
        if self.action == 'create':
            return CadastroImovelSerializer
        elif self.action == 'partial_update':
            return UpdateImovelSerializer
        else:
            return ListaImoveisSeriliazer

    def _filtrar_relatorio_por_status(self, request):
        status_em_analise = ['AGUARDANDO_ANALISE_PREVIA_SME', 'ENVIADO_COMAPRE',
                             'AGENDAMENTO_DA_VISTORIA', 'AGUARDANDO_RELATORIO_DE_VISTORIA',
                             'AGUARDANDO_LAUDO_DE_VALOR_LOCATICIO', 'SOLICITACAO_REALIZADA',
                             'RELATORIO_VISTORIA', 'LAUDO_VALOR_LOCATICIO']
        status_aprovados_vistoria = ['VISTORIA_APROVADA', 'ENVIADO_DRE',
                                     'FINALIZADO_APROVADO']
        status_reprovados_vistoria = ['VISTORIA_REPROVADA', 'FINALIZADO_REPROVADO']
        status_finalizados_reprovados = ['FINALIZADO_AREA_INSUFICIENTE',
                                         'FINALIZADO_DEMANDA_INSUFICIENTE',
                                         'FINALIZADO_NAO_ATENDE_NECESSIDADES']
        status_cancelados = ['CANCELADO']
        imoveis = Imovel.objects.filter(excluido=False).all()
        anos_selecionados = ""
        status_selecionados = ""
        status = []
        if (request.query_params.getlist('anos') != []):
            imoveis = imoveis.filter(criado_em__year__in=request.query_params.getlist('anos'))
            anos = request.query_params.getlist('anos')
            for idx, ano in enumerate(anos, 1):
                if idx != len(anos):
                    anos_selecionados = "{} {},".format(anos_selecionados, ano)
                else:
                    anos_selecionados = "{} {}".format(anos_selecionados, ano)
        total = imoveis.count()
        em_analise = 0
        aprovados_na_vistoria = 0
        reprovados_na_vistoria = 0
        finalizados_reprovados = 0
        cancelados = 0
        legenda = {
          'em_analise': False,
          'aprovados_na_vistoria': False,
          'reprovados_na_vistoria': False,
          'finalizados_reprovados': False,
          'cancelados': False
        }

        if(request.query_params.getlist('status') == []):
            em_analise = imoveis.filter(status__in=status_em_analise).count()
            aprovados_na_vistoria = imoveis.filter(status__in=status_aprovados_vistoria).count()
            reprovados_na_vistoria = imoveis.filter(status__in=status_reprovados_vistoria).count()
            finalizados_reprovados = imoveis.filter(status__in=status_finalizados_reprovados).count()
            cancelados = imoveis.filter(status__in=status_cancelados).count()
        else:
            if '1' in request.query_params.getlist('status'):
                status.append('EM ANÁLISE')
                em_analise = imoveis.filter(status__in=status_em_analise).count()
                legenda['em_analise'] = True
            if '2' in request.query_params.getlist('status'):
                status.append('VISTORIA APROVADA')
                aprovados_na_vistoria = imoveis.filter(status__in=status_aprovados_vistoria).count()
                legenda['aprovados_na_vistoria'] = True
            if '3' in request.query_params.getlist('status'):
                status.append('VISTORIA REPROVADA')
                reprovados_na_vistoria = imoveis.filter(status__in=status_reprovados_vistoria).count()
                legenda['reprovados_na_vistoria'] = True
            if '4' in request.query_params.getlist('status'):
                status.append('FINALIZADOS REPROVADOS')
                finalizados_reprovados = imoveis.filter(status__in=status_finalizados_reprovados).count()
                legenda['finalizados_reprovados'] = True
            if '5' in request.query_params.getlist('status'):
                status.append('CANCELADOS')
                cancelados = imoveis.filter(status__in=status_cancelados).count()
                legenda['cancelados'] = True

            status_selecionados = ", ".join(status)

        data = {'total': total,
                'em_analise': em_analise,
                'aprovados_na_vistoria': aprovados_na_vistoria,
                'reprovados_na_vistoria': reprovados_na_vistoria,
                'finalizados_reprovados': finalizados_reprovados,
                'cancelados': cancelados,
                'nome': request.user.first_name,
                'sobrenome': request.user.last_name,
                'rf': request.user.username,
                'anos_selecionados': anos_selecionados,
                'status_selecionados': status_selecionados,
                'data_hoje': datetime.datetime.strftime(datetime.datetime.now(), "%d/%m/%Y"),
                'legenda': legenda}

        return data

    def _filtrar_relatorio_demanda_territorial(self, request):
        imoveis = Imovel.objects.filter(excluido=False).annotate(demandaimovel__total=Sum('demandaimovel__bercario_i') + Sum('demandaimovel__bercario_ii') + Sum('demandaimovel__mini_grupo_i') + Sum('demandaimovel__mini_grupo_ii'))
        if (request.query_params.getlist('anos') != []):
            imoveis = imoveis.filter(criado_em__year__in=request.query_params.getlist('anos'))
        if request.query_params.getlist('setores') != []:
            imoveis = imoveis.filter(setor__codigo__in=request.query_params.getlist('setores'))
        if request.query_params.getlist('distritos') != []:
            imoveis = imoveis.filter(setor__distrito__id__in=request.query_params.getlist('distritos'))
        if request.query_params.get('dres') not in ['todas', None]:
            imoveis = imoveis.filter(setor__distrito__subprefeitura__dre__id=request.query_params.get('dres'))
        if request.query_params.getlist('demandas') != []:
            expressoes = {"1": Q(demandaimovel__total__lt=40), "2": Q(demandaimovel__total__gte=40, demandaimovel__total__lte=100), "3": Q(demandaimovel__total__gt=100)}
            expressao = None
            for d in request.query_params.getlist('demandas'):
                if expressao:
                    expressao = expressao | expressoes[d]
                else:
                    expressao = expressoes[d]
            imoveis = imoveis.filter(expressao)
        return imoveis

    def _filtrar_relatorio_area_construida(self, request):
        imoveis = Imovel.objects.all()
        resultado = {
            'ate_200': 0,
            '200_a_500': 0,
            'maior_500': 0,
        }
        if request.query_params.get('ano') not in ['todos', None]:
            imoveis = imoveis.filter(criado_em__year=request.query_params.get('ano'))
        if '1' in request.query_params.getlist('areas') or request.query_params.getlist('areas') == []:
            resultado['ate_200'] = imoveis.filter(area_construida__lt=200).count()
        if '2' in request.query_params.getlist('areas') or request.query_params.getlist('areas') == []:
            resultado['200_a_500'] = imoveis.filter(area_construida__gte=200, area_construida__lte=500).count()
        if '3' in request.query_params.getlist('areas') or request.query_params.getlist('areas') == []:
            resultado['maior_500'] = imoveis.filter(area_construida__gt=500).count()
        resultado['total_filtrado'] = (resultado['ate_200'] + resultado['200_a_500'] + resultado['maior_500'] )
        return resultado

    def _get_resultado_por_dre(self, imoveis, dres, todas_demandas, request):
        resultado_por_dre = {}
        for dre in dres:
            imoveis_filtrados = imoveis.filter(setor__distrito__subprefeitura__dre__nome=dre.nome)
            if todas_demandas:
                demanda_1 = imoveis_filtrados.filter(demandaimovel__total__lt=40).count()
                demanda_2 = imoveis_filtrados.filter(demandaimovel__total__gte=40, demandaimovel__total__lte=100).count()
                demanda_3 = imoveis_filtrados.filter(demandaimovel__total__gt=100).count()
                resultado_por_dre[f"{dre.nome}"] = {'demanda_1': demanda_1, 'demanda_2': demanda_2, 'demanda_3': demanda_3}
                if request.query_params.get('dres') not in ["todas", None]:
                    if resultado_por_dre[dre.nome] == {'demanda_1': 0, 'demanda_2': 0, 'demanda_3': 0,}:
                        del resultado_por_dre[dre.nome]
            else:
                resultado_por_dre[f"{dre.nome}"] = imoveis_filtrados.count()
                if request.query_params.get('dres') not in ["todas", None] :
                    if resultado_por_dre[dre.nome] == 0:
                        del resultado_por_dre[dre.nome]
        return resultado_por_dre

    def _get_resultado_por_distrito(self, imoveis, dres, todas_demandas):
        resultado_por_distrito = {}
        for dre in dres:
            imoveis_filtrados = imoveis.filter(setor__distrito__subprefeitura__dre__nome=dre.nome)
            resultado_por_distrito[f"{dre.nome}"] = {}
            if todas_demandas:
                for imovel in imoveis_filtrados:
                    demanda_1 = imoveis_filtrados.filter(demandaimovel__total__lt=40,
                                                         setor__distrito__id=imovel.setor.distrito.id).count()
                    demanda_2 = imoveis_filtrados.filter(demandaimovel__total__gte=40,
                                                         demandaimovel__total__lte=100,
                                                         setor__distrito__id=imovel.setor.distrito.id).count()
                    demanda_3 = imoveis_filtrados.filter(demandaimovel__total__gt=100,
                                                         setor__distrito__id=imovel.setor.distrito.id).count()
                    resultado_por_distrito[f"{dre.nome}"][f"{imovel.setor.distrito.nome}"] = {'demanda_1': demanda_1,
                                                                                              'demanda_2': demanda_2,
                                                                                              'demanda_3': demanda_3}
                if resultado_por_distrito[dre.nome] == {}:
                    del resultado_por_distrito[dre.nome]
            else:
                for imovel in imoveis_filtrados:
                    if resultado_por_distrito[f"{dre.nome}"].get(f"{imovel.setor.distrito.nome}") == None:
                        resultado_por_distrito[f"{dre.nome}"][f"{imovel.setor.distrito.nome}"] = 1
                    else:
                        count_distrito = resultado_por_distrito[f"{dre.nome}"][f"{imovel.setor.distrito.nome}"] + 1
                        resultado_por_distrito[f"{dre.nome}"][f"{imovel.setor.distrito.nome}"] = count_distrito
                if resultado_por_distrito[dre.nome] == {}:
                    del resultado_por_distrito[dre.nome]
        return resultado_por_distrito

    def _get_resultado_por_setor(self, imoveis, dres, todas_demandas):
        resultado_por_setor = {}
        for dre in dres:
            imoveis_filtrados = imoveis.filter(setor__distrito__subprefeitura__dre__nome=dre.nome)
            resultado_por_setor[f"{dre.nome}"] = {}
            if todas_demandas:
                for imovel in imoveis_filtrados:
                    demanda_1 = imoveis_filtrados.filter(demandaimovel__total__lt=40,
                                                         setor__distrito__id=imovel.setor.distrito.id).count()
                    demanda_2 = imoveis_filtrados.filter(demandaimovel__total__gte=40,
                                                         demandaimovel__total__lte=100,
                                                         setor__distrito__id=imovel.setor.distrito.id).count()
                    demanda_3 = imoveis_filtrados.filter(demandaimovel__total__gt=100,
                                                         setor__distrito__id=imovel.setor.distrito.id).count()
                    resultado = {'demanda_1': demanda_1,
                                 'demanda_2': demanda_2,
                                 'demanda_3': demanda_3}
                    resultado_por_setor[f"{dre.nome}"][f"{imovel.setor.distrito.nome}"] = {f"{imovel.setor.codigo}": resultado}
                if resultado_por_setor[dre.nome] == {}:
                    del resultado_por_setor[dre.nome]
            else:
                for imovel in imoveis_filtrados:
                    count_setor = Imovel.objects.filter(setor__codigo=imovel.setor.codigo).count()
                    resultado_por_setor[f"{dre.nome}"][f"{imovel.setor.distrito.nome}"] = {f"{imovel.setor.codigo}": count_setor}
                if resultado_por_setor[dre.nome] == {}:
                    del resultado_por_setor[dre.nome]
        return resultado_por_setor

    def _formatar_header(self, request, data):
        anos = request.query_params.getlist('anos')
        anos_selecionados = ""
        distritos_selecionados = ""
        setores_selecionados = ""
        areas_selecionadas = ""
        for idx, ano in enumerate(anos, 1):
            if idx != len(anos):
                anos_selecionados = "{} {},".format(anos_selecionados, ano)
            else:
                anos_selecionados = "{} {}".format(anos_selecionados, ano)

        if anos_selecionados == "":
            anos_selecionados = ' - '
        if request.query_params.get('dres') in [None, 'todas']:
            dre = 'Todas'
        else:
            dre = DiretoriaRegional.objects.filter(id=request.query_params.get('dres')).first().nome
        if request.query_params.getlist('demandas') == [] or len(request.query_params.getlist('demandas')) == 3 :
            demadas_selecionadas = 'Todas'
        else:
            demandas = request.query_params.getlist('demandas')
            demadas_selecionadas = ""
            for idx, demanda in enumerate(demandas, 1):
                if demanda == '1':
                    if idx != len(demandas):
                        demadas_selecionadas = "{} {},".format(demadas_selecionadas, "Baixa" )
                    else:
                        demadas_selecionadas = "{} {}".format(demadas_selecionadas, "Baixa")
                if demanda == '2':
                    if idx != len(demandas):
                        demadas_selecionadas = "{} {},".format(demadas_selecionadas, "Média" )
                    else:
                        demadas_selecionadas = "{} {}".format(demadas_selecionadas, "Média")
                if demanda == '3':
                    if idx != len(demandas):
                        demadas_selecionadas = "{} {},".format(demadas_selecionadas, "Alta" )
                    else:
                        demadas_selecionadas = "{} {}".format(demadas_selecionadas, "Alta")

        if request.query_params.getlist('distritos') != []:
            distritos = Distrito.objects.filter(id__in=request.query_params.getlist('distritos'))
            for idx, distrito in enumerate(distritos, 1):
                if idx != len(distritos):
                    distritos_selecionados = "{} {},".format(distritos_selecionados, distrito.nome)
                else:
                    distritos_selecionados = "{} {}".format(distritos_selecionados, distrito.nome)
        else:
            distritos_selecionados = " - "

        if request.query_params.getlist('setores') != []:
            setores = Setor.objects.filter(codigo__in=request.query_params.getlist('setores'))
            for idx, setor in enumerate(setores, 1):
                if idx != len(setores):
                    setores_selecionados = "{} {},".format(setores_selecionados, setor.codigo)
                else:
                    setores_selecionados = "{} {}".format(setores_selecionados, setor.codigo)
        else:
            setores_selecionados = " - "

        if ((request.query_params.getlist('areas') != []) and (len(request.query_params.getlist('areas')) != 3)):
            areas = request.query_params.getlist('areas')
            areas_selecionadas = ""
            for idx, area in enumerate(areas, 1):
                if area == '1':
                    if idx != len(areas):
                        areas_selecionadas = "{} {},".format(areas_selecionadas, "Abaixo de 200m²" )
                    else:
                        areas_selecionadas = "{} {}".format(areas_selecionadas, "Abaixo de 200m²")
                if area == '2':
                    if idx != len(areas):
                        areas_selecionadas = "{} {},".format(areas_selecionadas, "Entre 200m² e 500m²" )
                    else:
                        areas_selecionadas = "{} {}".format(areas_selecionadas, "Entre 200m² e 500m²")
                if area == '3':
                    if idx != len(areas):
                        areas_selecionadas = "{} {},".format(areas_selecionadas, "Acima de 500m²" )
                    else:
                        areas_selecionadas = "{} {}".format(areas_selecionadas, "Acima de 500m²")
        else:
            areas_selecionadas = "Todas"

        header = {
            'data_hoje': datetime.datetime.strftime(datetime.datetime.now(), "%d/%m/%Y"),
            'nome': request.user.first_name,
            'sobrenome': request.user.last_name,
            'rf': request.user.username,
            'anos_selecionados': anos_selecionados,
            'dres_selecionadas': dre,
            'demandas_selecionadas': demadas_selecionadas,
            'distritos_selecionados': distritos_selecionados,
            'setores_selecionados': setores_selecionados,
            'areas_selecionadas': areas_selecionadas,
        }
        resultado = {'header': header, 'data': data }
        return resultado

    def _filtrar_cadastros(self, request):
        queryset = Imovel.objects.filter(excluido=False).annotate(demandaimovel__total=Sum('demandaimovel__bercario_i') + Sum('demandaimovel__bercario_ii') + Sum('demandaimovel__mini_grupo_i') + Sum('demandaimovel__mini_grupo_ii'))
        if 'protocolo' in request.query_params:
            queryset = queryset.filter(id=request.query_params.get('protocolo'))
        if 'endereco' in request.query_params:
            queryset = queryset.filter(Q(endereco__icontains=request.query_params.get('endereco')))
        if 'area' in request.query_params:
            area = request.query_params.get('area')
            if area == '1':
                queryset = queryset.filter(area_construida__lt=200)
            if area == '2':
                queryset = queryset.filter(area_construida__gte=200, area_construida__lte=500)
            if area == '3':
                queryset = queryset.filter(area_construida__gt=500)
        if 'setor' in request.query_params:
            queryset = queryset.filter(setor__codigo=request.query_params.get('setor'))
        if 'distrito' in request.query_params:
            queryset = queryset.filter(setor__distrito__id=request.query_params.get('distrito'))
        if 'dre' in request.query_params:
            queryset = queryset.filter(setor__distrito__subprefeitura__dre__id=request.query_params.get('dre'))
        if 'status' in request.query_params:
            queryset = queryset.filter(status=request.query_params.get('status'))
        if 'demanda' in request.query_params:
            demanda = request.query_params.get('demanda')
            if demanda == '1':
                queryset = queryset.filter(demandaimovel__total__lt=40)
            if demanda == '2':
                queryset = queryset.filter(demandaimovel__total__gte=40, demandaimovel__total__lte=100)
            if demanda == '3':
                queryset = queryset.filter(demandaimovel__total__gt=100)
        if ('data_inicio' in request.query_params) and ('data_fim' in request.query_params):
            data_inicio = request.query_params.get('data_inicio').split('-')
            data_fim = request.query_params.get('data_fim').split('-')
            dates = [
                        datetime.date(int(data_inicio[0]), int(data_inicio[1]),  int(data_inicio[2])),
                        datetime.date(int(data_fim[0]), int(data_fim[1]),  int(data_fim[2])),
                    ]
            queryset = queryset.filter(criado_em__gte=dates[0], criado_em__lte=dates[1])
        return queryset

    def _gerar_planilha(self, imoveis):
        cabecalho = ['Número do Protocolo', 'Logradouro', 'DRE', 'Distrito',
                     'Setor', 'Número do IPTU', 'Área por m²',
                     'Proprietário/Representante', 'Telefone', 'Celular', 'E-mail',
                     'Status']

        count_fields = len(cabecalho)
        count_data = len(imoveis)
        wb = Workbook()
        ws = wb.active
        ws.title = "Relatório por Status"
        for ind, title in enumerate(cabecalho, 1):
            celula = ws.cell(row=1, column=ind)
            celula.value = title
            ws.column_dimensions[celula.column_letter].width = 20
            if ind in [1, 6]:
                ws.column_dimensions[celula.column_letter].width = 25
            if ind in [11, 12]:
                ws.column_dimensions[celula.column_letter].width = 50
            if ind in [2, 8]:
                ws.column_dimensions[celula.column_letter].width = 60

        for i, imovel in enumerate(imoveis, 2):
            ws.cell(row=i, column=1, value=imovel.protocolo)
            ws.cell(row=i, column=2, value=("%s, %s"%(imovel.endereco, imovel.numero)))
            if(imovel.setor != None):
                ws.cell(row=i, column=3, value=imovel.setor.distrito.subprefeitura.dre.first().sigla)
                ws.cell(row=i, column=4, value=imovel.setor.distrito.nome)
                ws.cell(row=i, column=5, value=imovel.setor.codigo)
            ws.cell(row=i, column=6, value=imovel.numero_iptu)
            ws.cell(row=i, column=7, value=imovel.area_construida)
            if(imovel.proponente != None):
                ws.cell(row=i, column=8, value=imovel.proponente.nome)
                ws.cell(row=i, column=9, value=imovel.proponente.telefone)
                ws.cell(row=i, column=10, value=imovel.proponente.celular)
                ws.cell(row=i, column=11, value=imovel.proponente.email)
            ws.cell(row=i, column=12, value=imovel.status.title)

        for linha in range(1, (count_data + 2)):
            for coluna in range(1, (count_fields + 1)):
                celula = ws.cell(row=linha, column=coluna)
                celula.font = Font(color="404040", size="12", bold=True)
                celula.border = Border(right=Side(border_style='thin', color='24292E'), left=Side(border_style='thin', color='24292E'), top=Side(border_style='thin', color='24292E'), bottom=Side(border_style='thin', color='24292E'))
                if linha == 1:
                    celula.fill = PatternFill(fill_type='solid', fgColor='8EAADC')
                elif(linha % 2) == 0:
                    celula.fill = PatternFill(fill_type='solid', fgColor='C5C5C5')
                else:
                    celula.fill = PatternFill(fill_type='solid', fgColor='EAEAEA')
                celula.alignment = Alignment(horizontal='center', vertical='center')

        result = BytesIO(save_virtual_workbook(wb))
        return result

    def _agrupa_por_mes_por_solicitacao(self, query_set: list) -> dict:
        # TODO: melhorar performance
        sumario = {'novos_cadastros': 0, 'proximos_ao_vencimento': 0, 'atrasados': 0}  # type: dict
        _25_dias_atras = datetime.date.today() - datetime.timedelta(days=25)
        _30_dias_atras = datetime.date.today() - datetime.timedelta(days=30)
        sumario['novos_cadastros'] = query_set.filter(criado_em__gte=_25_dias_atras).count()
        sumario['proximos_ao_vencimento'] = query_set.filter(criado_em__lt=_25_dias_atras,
                                                             criado_em__gte=_30_dias_atras).count()
        sumario['atrasados'] = query_set.filter(criado_em__lt=_30_dias_atras).count()
        return sumario

    @action(
        detail=False,
        methods=['GET'],
        url_path=f'ultimos-30-dias',
        permission_classes=(IsAuthenticated,))
    def ultimos_30_dias(self, request):
        query_set = Imovel.objects.filter(excluido=False, status=Imovel.workflow_class.initial_state).all()
        resumo_do_mes = self._agrupa_por_mes_por_solicitacao(query_set=query_set)
        return Response(resumo_do_mes, status=status.HTTP_200_OK)

    @action(
        detail=False,
        methods=['GET'],
        url_path=f'imoveis/novos-cadastros',
        permission_classes=(IsAuthenticated,))
    def imoveis_novos_cadastros(self, request):
        query_set = Imovel.objects.filter(criado_em__gt=datetime.date.today() - datetime.timedelta(days=25), excluido=False, status=Imovel.workflow_class.initial_state)
        page = self.paginate_queryset(query_set)
        serializer = self.get_serializer(page, many=True)
        return self.get_paginated_response(serializer.data)

    @action(
        detail=False,
        methods=['GET'],
        url_path=f'imoveis/proximos-ao-vencimento',
        permission_classes=(IsAuthenticated,))
    def imoveis_proximos_ao_vencimento(self, request):
        _25_dias_atras = datetime.date.today() - datetime.timedelta(days=25)
        _30_dias_atras = datetime.date.today() - datetime.timedelta(days=30)
        query_set = Imovel.objects.filter(criado_em__lt=_25_dias_atras,
                                          criado_em__gte=_30_dias_atras,
                                          excluido=False, status=Imovel.workflow_class.initial_state)
        page = self.paginate_queryset(query_set)
        serializer = self.get_serializer(page, many=True)
        return self.get_paginated_response(serializer.data)

    @action(
        detail=False,
        methods=['GET'],
        url_path=f'imoveis/atrasados',
        permission_classes=(IsAuthenticated,))
    def imoveis_atrasados(self, request):
        _30_dias_atras = datetime.date.today() - datetime.timedelta(days=30)
        query_set = Imovel.objects.filter(criado_em__lt=_30_dias_atras, excluido=False, status=Imovel.workflow_class.initial_state)
        page = self.paginate_queryset(query_set)
        serializer = self.get_serializer(page, many=True)
        return self.get_paginated_response(serializer.data)

    def retrieve(self, request, *args, **kwargs):
        serializer = self.get_serializer(self.get_object(), context={'request': request})
        return Response(serializer.data, status=status.HTTP_200_OK)

    def update(self, request, *args, **kwargs):
        if not request.user:
            return Response(status=status.HTTP_403_FORBIDDEN)
        return super().update(request, args, kwargs)

    def list(self, request, *args, **kwargs):
        queryset = self._filtrar_cadastros(request)
        page = self.paginate_queryset(queryset)
        serializer = self.get_serializer(page, many=True, context={'request': request})
        return self.get_paginated_response(serializer.data)

    @action(
        detail=False,
        methods=['PUT'],
        url_path=f'imoveis/update-status',
        permission_classes=(IsAuthenticated,))
    def update_status(self, request):
        imovel = Imovel.objects.filter(excluido=False).all().get(id=request.query_params.get('id'))
        imovel.situacao = request.query_params.get('situacao')
        imovel.escola = request.query_params.get('escola')
        imovel.codigo_eol = request.query_params.get('codigo_eol')
        imovel.save()
        serializer = self.get_serializer(imovel, context={'request': request})
        return Response(status=status.HTTP_200_OK, data=serializer.data)

    @action(
        detail=False,
        methods=['GET'],
        url_path=f'imoveis/exportar',
        permission_classes=(IsAuthenticated,)
        )
    def exportar(self, request):
        import time
        from django.db.models import Count, Q, Max
        from django.db.models.functions import Length
        from django.db import connection

        start_time = time.time()
        imoveis = self._filtrar_cadastros(request)
        count_anexos = 0
        count_fachada = 0
        count_interno = 0
        count_externo = 0
        count_iptu_itr = 0
        count_planta = 0

        tipo_0 = Count('plantafoto', filter=Q(plantafoto__tipo_documento__iexact='0'))
        tipo_1 = Count('plantafoto', filter=Q(plantafoto__tipo_documento__iexact='1'))
        tipo_2 = Count('plantafoto', filter=Q(plantafoto__tipo_documento__iexact='2'))
        tipo_3 = Count('plantafoto', filter=Q(plantafoto__tipo_documento__iexact='3'))
        tipo_4 = Count('plantafoto', filter=Q(plantafoto__tipo_documento__iexact='4'))

        for imovel in imoveis.annotate(tipo0=tipo_0).annotate(tipo1=tipo_1).annotate(tipo2=tipo_2).annotate(tipo3=tipo_3).annotate(tipo4=tipo_4):
            fachada = imovel.tipo0
            interno = imovel.tipo1
            externo = imovel.tipo2
            iptu_itr = imovel.tipo3
            planta = imovel.tipo4

            if count_fachada < fachada:
                count_fachada = fachada
            if count_interno < interno:
                count_interno = interno
            if count_externo < externo:
                count_externo = externo
            if count_iptu_itr < iptu_itr:
                count_iptu_itr = iptu_itr
            if count_planta < planta:
                count_planta = planta

        cabecalho = ['Protocolo', 'Data do Cadastro', 'Nome', 'E-mail', 'Celular', 'Telefone',
                        'CPF/CNPJ', 'CEP', 'Número do IPTU', 'Endereço', 'Bairro', 'Cidade',
                        'UF', 'Área Construída', 'DRE', 'Distrito', 'Setor', 'Status',
                        'Berçário I', 'Berçário II', 'Mini Grupo I', 'Mini Grupo II', 'Demanda Total']

        for i in range(1, (count_fachada+1)):
            cabecalho.append("Fotos da Fachada %i"%(i))

        for i in range(1, (count_interno+1)):
            cabecalho.append("Fotos do Ambiente Interno %i"%(i))

        for i in range(1, (count_externo+1)):
            cabecalho.append("Fotos de Área Externa %i"%(i))

        for i in range(1, (count_iptu_itr+1)):
            cabecalho.append("Cópia do IPTU ou ITR %i"%(i))

        for i in range(1, (count_planta+1)):
            cabecalho.append("Cópia da Planta ou Croqui %i"%(i))

        count_fields = len(cabecalho)
        count_data = imoveis.count()

        wb = Workbook()
        ws = wb.active
        ws.title = "Cadastros Realizados"

        border = Border(right=Side(border_style='thin', color='24292E'), left=Side(border_style='thin', color='24292E'), top=Side(border_style='thin', color='24292E'), bottom=Side(border_style='thin', color='24292E'))
        alignment = Alignment(horizontal='center', vertical='center')
        font = Font(color="404040", size="12", bold=True)

        fill_par = PatternFill(fill_type='solid', fgColor='C5C5C5')
        fill_impar = PatternFill(fill_type='solid', fgColor='EAEAEA')

        for ind, title in enumerate(cabecalho, 1):
            celula = ws.cell(row=1, column=ind)
            celula.value = title
            ws.column_dimensions[celula.column_letter].width = 40

        # Colorindo o cabeçalho
        for rows in ws.iter_rows(min_row=1, max_row=1, min_col=1):
            for cell in rows:
                cell.fill = PatternFill(fill_type='solid', fgColor='8EAADC')
                cell.font = Font(color="404040", size="12", bold=True)
                cell.border = Border(right=Side(border_style='thin', color='24292E'), left=Side(border_style='thin', color='24292E'), top=Side(border_style='thin', color='24292E'), bottom=Side(border_style='thin', color='24292E'))
                cell.alignment = Alignment(horizontal='center', vertical='center')

        name_length = Proponente.objects.values('nome').aggregate(name_length=Max(Length('nome')))['name_length']
        email_length = Proponente.objects.values('email').aggregate(email_length=Max(Length('email')))['email_length']
        endereco_length = Imovel.objects.values('endereco').aggregate(endereco_length=Max(Length('endereco')))['endereco_length']

        ws.column_dimensions["C"].width = name_length + 6
        ws.column_dimensions["D"].width = email_length + 8
        ws.column_dimensions["J"].width = endereco_length + 20

        def trata_celula(row, column, fill, value=None):
            celula = ws.cell(row=row, column=column, value=value)
            celula.fill = fill
            celula.font = font
            celula.border = border
            celula.alignment = alignment

            return celula

        for ind, imovel in enumerate(imoveis.select_related('proponente', 'setor', 'demandaimovel').prefetch_related("plantafoto_set"), 2):
            fill = fill_par if (ind % 2) == 0 else fill_impar

            trata_celula(row=ind, column=1, value=imovel.protocolo, fill=fill)
            trata_celula(row=ind, column=2, value=imovel.criado_em, fill=fill)
            if(imovel.proponente != None):
                trata_celula(row=ind, column=3, fill=fill, value=imovel.proponente.nome)
                trata_celula(row=ind, column=4, fill=fill, value=imovel.proponente.email)
                trata_celula(row=ind, column=5, fill=fill, value=imovel.proponente.celular)
                trata_celula(row=ind, column=6, fill=fill, value=imovel.proponente.telefone)
                trata_celula(row=ind, column=7, fill=fill, value=imovel.proponente.cpf_cnpj)
            trata_celula(row=ind, fill=fill, column=8, value=imovel.cep)
            trata_celula(row=ind, fill=fill, column=9, value=imovel.numero_iptu)
            trata_celula(row=ind, fill=fill, column=10, value=("%s, %s"%(imovel.endereco, imovel.numero)))
            trata_celula(row=ind, fill=fill, column=11, value=imovel.bairro)
            trata_celula(row=ind, fill=fill, column=12, value=imovel.cidade)
            trata_celula(row=ind, fill=fill, column=13, value=imovel.uf)
            trata_celula(row=ind, fill=fill, column=14, value=imovel.area_construida)
            if(imovel.setor != None):
                trata_celula(row=ind, column=15, fill=fill, value=imovel.setor.distrito.subprefeitura.dre.first().sigla)
                trata_celula(row=ind, column=16, fill=fill, value=imovel.setor.distrito.nome)
                trata_celula(row=ind, column=17, fill=fill, value=imovel.setor.codigo)
            if(imovel.status != None):
                trata_celula(row=ind, column=18, fill=fill, value=imovel.status.title)
            if(hasattr(imovel, 'demandaimovel')):
                trata_celula(row=ind, column=19, fill=fill, value=imovel.demandaimovel.bercario_i)
                trata_celula(row=ind, column=20, fill=fill, value=imovel.demandaimovel.bercario_ii)
                trata_celula(row=ind, column=21, fill=fill, value=imovel.demandaimovel.mini_grupo_i)
                trata_celula(row=ind, column=22, fill=fill, value=imovel.demandaimovel.mini_grupo_ii)
                trata_celula(row=ind, column=23, fill=fill, value=imovel.demandaimovel.total)
            fachada = 0
            interno = 0
            externo = 0
            iptu_itr = 0
            planta = 0
            for x, anexo in enumerate(imovel.anexos, 1):
                tipo = anexo.get_tipo_documento_display()
                if tipo == "Fotos da Fachada":
                    fachada = (fachada + 1)
                    celula = trata_celula(row=ind, fill=fill, column=(23 + fachada))

                if tipo == "Fotos do Ambiente Interno":
                    interno = (interno + 1)
                    celula = trata_celula(row=ind, fill=fill, column=(23 + count_fachada + interno))

                if tipo == "Fotos de Área Externa":
                    externo = (externo + 1)
                    celula = trata_celula(row=ind, fill=fill, column=(23 + count_fachada + count_interno +externo))

                if tipo == "Cópia do IPTU ou ITR":
                    iptu_itr = (iptu_itr + 1)
                    celula = trata_celula(row=ind, fill=fill, column=(23 + count_fachada + count_interno + count_externo + iptu_itr))

                if tipo == "Cópia da Planta ou Croqui":
                    planta = (planta + 1)
                    celula = trata_celula(row=ind, fill=fill, column=(23 + count_fachada + count_interno + count_externo + count_iptu_itr + planta))
                link = "%s%s"%(env.str("URL_HOSTNAME_WITHOUT_SLASH_API", default=""), anexo.arquivo.url)
                celula.value = '=HYPERLINK("{}", "{}")'.format(link, anexo.get_tipo_documento_display())


        print(f"Tempo total: {(time.time()-start_time)} em segundos")
        print(f"Quantidade de conexões {len(connection.queries)}")

        result = BytesIO(save_virtual_workbook(wb))
        filename = 'cadastros-realizados.xlsx'
        response = HttpResponse(
            result,
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = 'attachment; filename=%s' % filename
        return response

    @action(detail=False,
            methods=['get'],
            url_path='checa-iptu-ja-existe/(?P<numero_iptu>.*)')
    def checa_iptu_ja_existe(self, request, numero_iptu=None):
        iptu_existe = numero_iptu in Imovel.objects.filter(excluido=False).all().values_list('numero_iptu', flat=True)
        iptu_valido = checa_digito_verificador_iptu(numero_iptu)
        return Response(
            {'iptu_existe': iptu_existe, 'iptu_valido': iptu_valido}, status=status.HTTP_200_OK
        )

    @action(detail=False, methods=['post'], url_path='checa-endereco-imovel-ja-existe')
    def checa_endereco_imovel_ja_existe(self, request):
        endereco_existe = Imovel.objects.filter(
            cep=request.data.get('cep'),
            endereco=request.data.get('endereco'),
            bairro=request.data.get('bairro'),
            numero=request.data.get('numero'),
            excluido=False
        ).exists()
        return Response(
            {'endereco_existe': endereco_existe}, status=status.HTTP_200_OK
        )

    @action(detail=False,
            methods=['POST'],
            url_path='imoveis/envia-comapre',
            permission_classes=(IsAuthenticated,))
    def enviar_para_comapre(self, request):
        imovel = Imovel.objects.get(id=request.query_params.get('imovel'))
        user = request.user
        if (request.query_params.get('enviar_email') == 'true'):
            enviar_email = True
        else:
            enviar_email = False
        data_agendada = request.query_params.get('data_agendada')
        justificativa = ""
        if 'justificativa' in request.query_params:
            justificativa=request.query_params.get('justificativa')
        imovel.sme_analisa_previamente(user=user)
        imovel.envia_a_comapre(user=user, data_agendada=data_agendada, enviar_email=enviar_email, justificativa=justificativa)
        if (enviar_email):
            data = imovel.as_dict()
            template = "envia_a_comapre"
            subject = f"Assunto: Cadastro de imóvel – Protocolo nº {data['protocolo']} – Solicitação de vistoria."
            email = data['proponente_email']
            task_send_email_to_usuario.delay(subject, template, data, email)
        serializer = self.get_serializer(imovel, context={'request': request})
        return Response(status=status.HTTP_200_OK, data=serializer.data)

    @action(detail=False,
            methods=['POST'],
            url_path='imoveis/finaliza',
            permission_classes=(IsAuthenticated,))
    def finalizar(self, request):
        imovel = Imovel.objects.get(id=request.query_params.get('imovel'))
        user = request.user
        if (request.query_params.get('enviar_email') == 'true'):
            enviar_email = True
        else:
            enviar_email = False
        justificativa = ""
        if 'justificativa' in request.query_params:
            justificativa=request.query_params.get('justificativa')
        if 'resultado' in request.query_params:
            if(request.query_params.get('resultado') == '0'):
                imovel.sme_analisa_previamente(user=user, justificativa=justificativa)
                imovel.finaliza_area_insuficiente(user=user, enviar_email=enviar_email)
                if (enviar_email):
                    data = imovel.as_dict()
                    template = "finaliza_area_insuficiente"
                    subject = f"Assunto: Cadastro de imóvel – Protocolo nº {data['protocolo']} – Área Insuficiente."
                    email = data['proponente_email']
                    task_send_email_to_usuario.delay(subject, template, data, email)
            if(request.query_params.get('resultado') == '1'):
                imovel.sme_analisa_previamente(user=user, justificativa=justificativa)
                imovel.finaliza_demanda_insuficiente(user=user, enviar_email=enviar_email)
                if (enviar_email):
                    data = imovel.as_dict()
                    template = "finalizado_demanda_insuficiente"
                    subject = f"Assunto: Cadastro de imóvel – Protocolo nº {data['protocolo']} – Demanda Insuficiente."
                    email = data['proponente_email']
                    task_send_email_to_usuario.delay(subject, template, data, email)
            if(request.query_params.get('resultado') == '2'):
                imovel.sme_analisa_previamente(user=user, justificativa=justificativa)
                imovel.finaliza_nao_atende_necessidades(user=user, enviar_email=enviar_email)
                if (enviar_email):
                    data = imovel.as_dict()
                    template = "finaliza_nao_atende_necessidades"
                    subject = f"Assunto: Cadastro de imóvel – Protocolo nº {data['protocolo']} – Não atende às necessidades da SME."
                    email = data['proponente_email']
                    task_send_email_to_usuario.delay(subject, template, data, email)
            if(request.query_params.get('resultado') == '3'):
                imovel.finaliza_aprovado(user=user, enviar_email=enviar_email, justificativa=justificativa)
            if(request.query_params.get('resultado') == '4'):
                imovel.finaliza_reprovado(user=user, enviar_email=enviar_email, justificativa=justificativa)
        serializer = self.get_serializer(imovel, context={'request': request})
        return Response(status=status.HTTP_200_OK, data=serializer.data)

    @action(detail=False,
            methods=['POST'],
            url_path='imoveis/agenda-vistoria',
            permission_classes=(IsAuthenticated,))
    def agendar_vistoria(self, request):
        imovel = Imovel.objects.get(id=request.query_params.get('imovel'))
        user = request.user
        if (request.query_params.get('enviar_email') == 'true'):
            enviar_email = True
        else:
            enviar_email = False
        data_agendada = request.query_params.get('data_agendada')
        imovel.agenda_vistoria(user=user, data_agendada=data_agendada, enviar_email=enviar_email)
        if (enviar_email):
            data = imovel.as_dict()
            template = "agenda_vistoria"
            subject = f"Assunto: Cadastro de imóvel – Protocolo nº {data['protocolo']} –  Agendamento de vistoria."
            email = data['proponente_email']
            task_send_email_to_usuario.delay(subject, template, data, email)
        imovel.aguarda_relatorio_vistoria(user=user)
        serializer = self.get_serializer(imovel, context={'request': request})
        return Response(status=status.HTTP_200_OK, data=serializer.data)

    @action(detail=False,
            methods=['POST'],
            url_path='imoveis/relatorio-vistoria',
            permission_classes=(IsAuthenticated,))
    def relatorio_vistoria(self, request):
        imovel = Imovel.objects.get(id=request.query_params.get('imovel'))
        user = request.user
        imovel.relatorio_vistoria(user=user)
        imovel.aguarda_laudo_valor_locaticio(user=user)
        serializer = self.get_serializer(imovel, context={'request': request})
        return Response(status=status.HTTP_200_OK, data=serializer.data)

    @action(detail=False,
            methods=['POST'],
            url_path='imoveis/laudo-locaticio',
            permission_classes=(IsAuthenticated,))
    def laudo_locaticio(self, request):
        imovel = Imovel.objects.get(id=request.query_params.get('imovel'))
        user = request.user
        imovel.laudo_valor_locaticio(user=user)
        serializer = self.get_serializer(imovel, context={'request': request})
        return Response(status=status.HTTP_200_OK, data=serializer.data)

    @action(detail=False,
            methods=['POST'],
            url_path='imoveis/resultado-vistoria',
            permission_classes=(IsAuthenticated,))
    def resultado_vistoria(self, request):
        imovel = Imovel.objects.get(id=request.query_params.get('imovel'))
        user = request.user
        if (request.query_params.get('enviar_email') == 'true'):
            enviar_email = True
        else:
            enviar_email = False
        if (request.query_params.get('resultado_da_vistoria') == '0'):
            imovel.aprova_vistoria(user=user, enviar_email=enviar_email)
            if (enviar_email):
                data = imovel.as_dict()
                template = "aprova_vistoria"
                subject = f"Assunto: Cadastro de imóvel – Protocolo nº {data['protocolo']} –  Vistoria Aprovada."
                email = data['proponente_email']
                task_send_email_to_usuario.delay(subject, template, data, email)
        else:
            imovel.reprova_vistoria(user=user, enviar_email=enviar_email)
            if (enviar_email):
                data = imovel.as_dict()
                template = "reprova_vistoria"
                subject = f"Assunto: Cadastro de imóvel – Protocolo nº {data['protocolo']} –  Vistoria Reprovada."
                email = data['proponente_email']
                task_send_email_to_usuario.delay(subject, template, data, email)
        serializer = self.get_serializer(imovel, context={'request': request})
        return Response(status=status.HTTP_200_OK, data=serializer.data)

    @action(detail=False,
            methods=['POST'],
            url_path='imoveis/envia-dre',
            permission_classes=(IsAuthenticated,))
    def enviar_para_dre(self, request):
        imovel = Imovel.objects.get(id=request.query_params.get('imovel'))
        user = request.user
        if (request.query_params.get('enviar_email') == 'true'):
            enviar_email = True
        else:
            enviar_email = False
        data_agendada = request.query_params.get('data_agendada')
        processo_sei = request.query_params.get('processo_sei')
        nome_da_unidade = request.query_params.get('nome_da_unidade')
        imovel.envia_a_dre(user=user, enviar_email=enviar_email,
                           processo_sei=processo_sei, nome_da_unidade=nome_da_unidade,
                           data_agendada=data_agendada)
        if (enviar_email):
            data = imovel.as_dict()
            template = "envia_dre"
            subject = f"Assunto: Cadastro de imóvel – Protocolo nº {data['protocolo']} –  Encaminhado à Diretoria Regional de Educação {data['diretoria_regional_educacao']}."
            email = data['proponente_email']
            task_send_email_to_usuario.delay(subject, template, data, email)
        serializer = self.get_serializer(imovel, context={'request': request})
        return Response(status=status.HTTP_200_OK, data=serializer.data)

    @action(detail=False,
            methods=['POST'],
            url_path='imoveis/cancela',
            permission_classes=(IsAuthenticated,))
    def cancelar(self, request):
        imovel = Imovel.objects.get(id=request.query_params.get('imovel'))
        user = request.user
        data_agendada = datetime.datetime.now()
        imovel.cancela(user=user, data_agendada=data_agendada)
        data = imovel.as_dict()
        template = "cancela_cadastro"
        subject = f"Assunto: Cadastro de imóvel – Protocolo nº {data['protocolo']} – Protocolo Cancelado."
        email = data['proponente_email']
        task_send_email_to_usuario.delay(subject, template, data, email)
        serializer = self.get_serializer(imovel, context={'request': request})
        return Response(status=status.HTTP_200_OK, data=serializer.data)

    @action(detail=False,
            methods=['POST'],
            url_path='imoveis/reativa',
            permission_classes=(IsAuthenticated,))
    def reativar(self, request):
        imovel = Imovel.objects.get(id=request.query_params.get('imovel'))
        user = request.user
        imovel.reativa(user=user)
        serializer = self.get_serializer(imovel, context={'request': request})
        return Response(status=status.HTTP_200_OK, data=serializer.data)

    @action(detail=False,
            methods=['get'],
            url_path='imoveis/get-as-dict')
    def get_as_dict(self, request):
        imovel = Imovel.objects.get(id=request.query_params.get('imovel'))
        return Response(status=status.HTTP_200_OK, data=imovel.as_dict())

    @action(detail=False,
            methods=['get'],
            url_path='imoveis/anos')
    def anos_imoveis(self, request):
        imoveis = Imovel.objects.filter(excluido=False).all()
        anos = []
        for imovel in imoveis:
            ano = datetime.datetime.strftime(imovel.criado_em, "%Y")
            if ano not in anos:
                anos.append(ano)
        return Response(status=status.HTTP_200_OK, data=anos)

    @action(detail=False,
            methods=['get'],
            url_path='imoveis/filtrar-por-status')
    def filtrar_por_status(self, request):
        data = self._filtrar_relatorio_por_status(request)
        return Response(status=status.HTTP_200_OK, data=data)

    @action(detail=False,
            methods=['get'],
            url_path='imoveis/filtrar-por-demanda-territorial')
    def filtrar_demanda_territorial(self, request):
        data = self._filtrar_relatorio_demanda_territorial(request)
        dres = DiretoriaRegional.objects.all()
        total = Imovel.objects.filter(excluido=False).all().count()
        todas_demandas = (len(request.query_params.getlist('demandas')) == 3)
        if request.query_params.get('tipo_resultado') == 'dre':
            data = self._get_resultado_por_dre(data, dres, todas_demandas, request)
        if request.query_params.get('tipo_resultado') == 'distrito':
            data = self._get_resultado_por_distrito(data, dres, todas_demandas)
        if request.query_params.get('tipo_resultado') == 'setor':
            data = self._get_resultado_por_setor(data, dres, todas_demandas)
        return Response(status=status.HTTP_200_OK, data={'data': data, 'total': total})

    @action(detail=False,
            methods=['get'],
            url_path='imoveis/filtrar-por-area-construida')
    def filtrar_area_construida(self, request):
        data = self._filtrar_relatorio_area_construida(request)
        data['total_imoveis'] = Imovel.objects.filter(excluido=False).all().count()
        return Response(status=status.HTTP_200_OK, data=data)

    @action(detail=False,
            methods=['get'],
            url_path='imoveis/relatorio-area-construida-xls')

    def relatorio_area_construida_xls(self, request):
        imoveis = None
        imoveis_1 = Imovel.objects.filter(area_construida__lt=200)
        imoveis_2 = Imovel.objects.filter(area_construida__gte=200, area_construida__lte=500)
        imoveis_3 = Imovel.objects.filter(area_construida__gt=500)
        if '1' in request.query_params.getlist('areas') or request.query_params.getlist('areas') == []:
            imoveis = imoveis_1
        if '2' in request.query_params.getlist('areas') or request.query_params.getlist('areas') == []:
            try:
                imoveis = imoveis | imoveis_2
            except:
                imoveis = imoveis_2
        if '3' in request.query_params.getlist('areas') or request.query_params.getlist('areas') == []:
            try:
                imoveis = imoveis | imoveis_3
            except:
                imoveis = imoveis_3
        if request.query_params.get('ano') not in ['todos', None]:
            try:
                imoveis = imoveis.filter(criado_em__year=request.query_params.get('ano'))
            except:
                imoveis = Imovel.objects.filter(criado_em__year=request.query_params.get('ano'))
        if imoveis == None:
            imoveis = Imovel.objects.all()
        result = self._gerar_planilha(imoveis)
        filename = 'relatorio-por-area-construida.xlsx'
        response = HttpResponse(
            result,
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = 'attachment; filename=%s' % filename
        return response

    @action(detail=False,
            methods=['get'],
            url_path='imoveis/relatorio-por-status-xls')
    def relatorio_por_status_xls(self, request):
        status_em_analise = ['AGUARDANDO_ANALISE_PREVIA_SME', 'ENVIADO_COMAPRE',
                             'AGENDAMENTO_DA_VISTORIA', 'AGUARDANDO_RELATORIO_DE_VISTORIA',
                             'AGUARDANDO_LAUDO_DE_VALOR_LOCATICIO', 'SOLICITACAO_REALIZADA',
                             'RELATORIO_VISTORIA', 'LAUDO_VALOR_LOCATICIO']
        status_aprovados_vistoria = ['VISTORIA_APROVADA', 'ENVIADO_DRE',
                                     'FINALIZADO_APROVADO']
        status_reprovados_vistoria = ['VISTORIA_REPROVADA', 'FINALIZADO_REPROVADO']
        status_finalizados_reprovados = ['FINALIZADO_AREA_INSUFICIENTE',
                                         'FINALIZADO_DEMANDA_INSUFICIENTE',
                                         'FINALIZADO_NAO_ATENDE_NECESSIDADES']
        status_cancelados = ['CANCELADO']
        imoveis = Imovel.objects.filter(excluido=False).all()
        em_analise = []
        aprovados_na_vistoria = []
        reprovados_na_vistoria = []
        finalizados_reprovados = []
        cancelados = []
        if (request.query_params.getlist('anos') != []):
            imoveis = imoveis.filter(criado_em__year__in=request.query_params.getlist('anos'))
        if(request.query_params.getlist('status') == []):
            em_analise = imoveis.filter(status__in=status_em_analise)
            aprovados_na_vistoria = imoveis.filter(status__in=status_aprovados_vistoria)
            reprovados_na_vistoria = imoveis.filter(status__in=status_reprovados_vistoria)
            finalizados_reprovados = imoveis.filter(status__in=status_finalizados_reprovados)
            cancelados = imoveis.filter(status__in=status_cancelados)
        else:
            if '1' in request.query_params.getlist('status'):
                em_analise = imoveis.filter(status__in=status_em_analise)
            if '2' in request.query_params.getlist('status'):
                aprovados_na_vistoria = imoveis.filter(status__in=status_aprovados_vistoria)
            if '3' in request.query_params.getlist('status'):
                reprovados_na_vistoria = imoveis.filter(status__in=status_reprovados_vistoria)
            if '4' in request.query_params.getlist('status'):
                finalizados_reprovados = imoveis.filter(status__in=status_finalizados_reprovados)
            if '5' in request.query_params.getlist('status'):
                cancelados = imoveis.filter(status__in=status_cancelados)
        imoveis = list(chain(em_analise, aprovados_na_vistoria))
        imoveis = list(chain(imoveis, reprovados_na_vistoria))
        imoveis = list(chain(imoveis, finalizados_reprovados))
        imoveis = list(chain(imoveis, cancelados))
        result = self._gerar_planilha(imoveis)
        filename = 'relatorio-por-status.xlsx'
        response = HttpResponse(
            result,
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = 'attachment; filename=%s' % filename
        return response

    @action(detail=False,
            methods=['get'],
            url_path='imoveis/relatorio-por-status-pdf')
    def relatorio_por_status_pdf(self, request, *args, **kwargs):
        data = self._filtrar_relatorio_por_status(request)

        html_string = render_to_string('imovel/relatorios/relatorio_por_status.html', {'data': data})
        html = HTML(string=html_string)

        html.write_pdf(target='/tmp/relatorio_por_status.pdf');

        fs = FileSystemStorage('/tmp')
        with fs.open('relatorio_por_status.pdf') as pdf:
            response = HttpResponse(pdf, content_type='application/pdf')
            response['Content-Disposition'] = 'attachment; filename="relatorio_por_status.pdf"'
            return response

        return response

    @action(detail=False,
            methods=['get'],
            url_path='imoveis/relatorio-por-demanda-xls')
    def relatorio_por_demanada_xls(self, request):
        imoveis = self._filtrar_relatorio_demanda_territorial(request)
        result = self._gerar_planilha(imoveis)
        filename = 'relatorio-por-demanda-territorial.xlsx'
        response = HttpResponse(
            result,
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = 'attachment; filename=%s' % filename
        return response

    @action(detail=False,
            methods=['get'],
            url_path='imoveis/relatorio-por-demanda-pdf')
    def relatorio_por_demanda_pdf(self, request, *args, **kwargs):
        imoveis = self._filtrar_relatorio_demanda_territorial(request)
        dres = DiretoriaRegional.objects.all()
        total = Imovel.objects.filter(excluido=False).all().count()
        todas_demandas = (len(request.query_params.getlist('demandas')) == 3)
        if request.query_params.get('tipo_resultado') == 'dre':
            data = self._get_resultado_por_dre(imoveis, dres, todas_demandas, request)
            resultado = self._formatar_header(request, data)
            html_string = render_to_string('imovel/relatorios/demanda_territorial_por_dre.html', resultado)
        if request.query_params.get('tipo_resultado') == 'distrito':
            data = self._get_resultado_por_distrito(imoveis, dres, todas_demandas)
            resultado = self._formatar_header(request, data)
            html_string = render_to_string('imovel/relatorios/demanda_territorial_por_distrito.html', resultado)
        if request.query_params.get('tipo_resultado') == 'setor':
            data = self._get_resultado_por_setor(imoveis, dres, todas_demandas)
            resultado = self._formatar_header(request, data)
            html_string = render_to_string('imovel/relatorios/demanda_territorial_por_setor.html', resultado)

        html = HTML(string=html_string)
        html.write_pdf(target='/tmp/relatorio_por_demanda.pdf');

        fs = FileSystemStorage('/tmp')
        with fs.open('relatorio_por_demanda.pdf') as pdf:
            response = HttpResponse(pdf, content_type='application/pdf')
            response['Content-Disposition'] = 'attachment; filename="relatorio_por_demanda.pdf"'
            return response

        return response

    @action(detail=False,
            methods=['get'],
            url_path='imoveis/relatorio-area-construida-pdf')
    def relatorio_area_construida_pdf(self, request, *args, **kwargs):
        data = self._filtrar_relatorio_area_construida(request)
        data['total_imoveis'] = Imovel.objects.filter(excluido=False).all().count()
        resultado = self._formatar_header(request, data)
        html_string = render_to_string('imovel/relatorios/relatorio_area_construida.html', resultado)

        html = HTML(string=html_string)
        html.write_pdf(target='/tmp/relatorio_area_construida.pdf');

        fs = FileSystemStorage('/tmp')
        with fs.open('relatorio_area_construida.pdf') as pdf:
            response = HttpResponse(pdf, content_type='application/pdf')
            response['Content-Disposition'] = 'attachment; filename="relatorio_area_construida.pdf"'
            return response

        return response

    @action(detail=False,
            methods=['get'],
            url_path='imoveis/relatorio-cadastro-pdf')
    def relatorio_cadastro_pdf(self, request, *args, **kwargs):
        data = Imovel.objects.get(id=request.query_params.get('id')).as_dict()
        data['solicitante_nome'] = f"{request.user.first_name} {request.user.last_name}"
        data['solicitante_rf'] = request.user.username

        from ..relatorio import relatorio_cadastro

        html_string = relatorio_cadastro(data)

        html = HTML(string=html_string, base_url=request.build_absolute_uri('/'))
        html.write_pdf(target='/tmp/relatorio_cadastro.pdf')

        fs = FileSystemStorage('/tmp')
        with fs.open('relatorio_cadastro.pdf') as pdf:
            response = HttpResponse(pdf, content_type='application/pdf')
            response['Content-Disposition'] = 'attachment; filename="relatorio_cadastro.pdf"'
            return response

        return response

class DemandaRegiao(APIView):
    """
    Encapsula a chamada a API de demanda
    """
    permission_classes = (AllowAny,)

    def get(self, request, param1, param2, format=None):
        url = f'{settings.SCIEDU_URL}/{param1}/{param2}'
        headers = {
            "Authorization": f'Token {settings.SCIEDU_TOKEN}',
            "Content-Type": "application/json"
        }
        response = requests.get(url, headers=headers)
        return Response(response.json(), status=response.status_code)


class AnexosViewset(mixins.CreateModelMixin,
                    mixins.DestroyModelMixin,
                    viewsets.GenericViewSet):
    permission_classes = [IsAuthenticated]
    lookup_field = 'uuid'
    queryset = PlantaFoto.objects.all()
    serializer_class = AnexoSerializer

    def get_serializer_class(self):
        if self.action == 'create':
            return AnexoCreateSerializer
        return AnexoSerializer
